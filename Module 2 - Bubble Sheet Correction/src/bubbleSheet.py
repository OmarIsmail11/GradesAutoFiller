import cv2 as cv
import numpy as np
from answersExtraction import *
from idExtraction import *
from paperExtraction import *
import pandas as pd

import cv2 as cv
import numpy as np
import pandas as pd
from answersExtraction import *
from idExtraction import *
from paperExtraction import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def readModelAnswer(filePath):
    model_answers = []
    with open(filePath, "r") as f:
        for line in f:
            answer = line.strip()
            if answer:
                model_answers.append(answer)
    return model_answers


def gradePapers(paperImages, modelAnswers):
    outputExcel="grades.xlsx"
    results = []

    maxGrade = len(modelAnswers)

    for imgPath in paperImages:
        paper = extractPaper(imgPath)
        studentID = detectID(paper)
        studentAnswers = detectAllAnswers(paper)

        grades = []
        for studentAns, correctAns in zip(studentAnswers, modelAnswers):
            if studentAns == 'Z':
                grade = 0
            elif studentAns == correctAns:
                grade = 1
            else:
                grade = -1
            grades.append(grade)

        total_score = sum(grades)
        results.append([studentID] + grades + [total_score])

    # Prepare DataFrame
    colNames = ["ID"] + [f"Q{i+1}" for i in range(len(modelAnswers))] + [f"Total / {maxGrade}"]
    df = pd.DataFrame(results, columns = colNames)

    # Save to Excel
    df.to_excel(outputExcel, index=False)

    # Load Excel for formatting
    wb = load_workbook(outputExcel)
    ws = wb.active

    yellow_fill = PatternFill(start_color = "FFFFFF00", end_color = "FFFFFF00", fill_type="solid")  # wrong (blank)
    red_fill = PatternFill(start_color = "FFFF0000", end_color = "FFFF0000", fill_type="solid")  # wrong (multiple)

    # Apply highlights
    for row_idx in range(2, len(df)+2):  # skip header
        for col_idx in range(2, 2 + len(modelAnswers)):  # Q1..Qn
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value == -1:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill
            elif cell_value == 0:
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(outputExcel)
    print(f"Graded Excel with highlights saved to {outputExcel}")

filePath = r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\modelAnswer.txt"
modelAnswers = readModelAnswer(filePath)

paperPaths = [
    r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\Module 2 - Bubble Sheet Correction\data\images\1.jpg",
    r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\Module 2 - Bubble Sheet Correction\data\images\2.jpg",
    r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\Module 2 - Bubble Sheet Correction\data\images\3.jpg",
    r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\Module 2 - Bubble Sheet Correction\data\images\4.jpg",
    r"D:\Omar\Image Processing & Computer Vision\GradesAutoFiller\Module 2 - Bubble Sheet Correction\data\images\5.jpg"
]

gradePapers(paperPaths, modelAnswers)


