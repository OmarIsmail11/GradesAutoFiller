import os
import sys

project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if project_root not in sys.path:
    sys.path.insert(0, project_root)


from Module2BubbleSheetCorrection.src.answersExtraction import *
from Module2BubbleSheetCorrection.src.idExtraction import *
from Module2BubbleSheetCorrection.src.paperExtraction import *
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def readModelAnswer(filePath):
    model_answers = []
    with open(filePath, "r") as f:
        for line in f:
            answer = line.strip()
            if answer:
                model_answers.append(answer)
    return model_answers


def gradePapers(paperImages, modelAnswers, outputExcel="grades.xlsx"):
    results = []

    maxGrade = len(modelAnswers)

    for imgPath in paperImages:
        paper = extractPaper(imgPath)
        studentID = detectID(paper)
        studentAnswers = detectAllAnswers(paper)

        grades = []
        for studentAns, correctAns in zip(studentAnswers, modelAnswers):
            if studentAns == 'Z':
                grade = 0
            elif studentAns == correctAns:
                grade = 1
            else:
                grade = -1
            grades.append(grade)

        total_score = sum(grades)
        results.append([studentID] + grades + [total_score])

    # Prepare DataFrame
    colNames = ["ID"] + [f"Q{i+1}" for i in range(len(modelAnswers))] + [f"Total / {maxGrade}"]
    df = pd.DataFrame(results, columns = colNames)

    # Save to Excel
    df.to_excel(outputExcel, index=False)

    # Load Excel for formatting
    wb = load_workbook(outputExcel)
    ws = wb.active

    yellow_fill = PatternFill(start_color = "FFFFFF00", end_color = "FFFFFF00", fill_type="solid")  # wrong (blank)
    red_fill = PatternFill(start_color = "FFFF0000", end_color = "FFFF0000", fill_type="solid")  # wrong (multiple)

    # Apply highlights
    for row_idx in range(2, len(df)+2):  # skip header
        for col_idx in range(2, 2 + len(modelAnswers)):  # Q1..Qn
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value == -1:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill
            elif cell_value == 0:
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(outputExcel)
    print(f"Graded Excel with highlights saved to {outputExcel}")
    
    return outputExcel



# filePath = r"C:/Users/youse/Desktop/University/Image/GradesAutoFiller/modelAnswer.txt"
# modelAnswers = readModelAnswer(filePath)

# paperPaths = [r"C:\Users\youse\Downloads\Bubble_sheet\zipped\1\ID2Q13CH3-20220106T190109Z-001\ID2Q13CH3\4dc7740c-2781-49c5-8f91-bb59b4d92bc8.jpg",
#               r"C:\Users\youse\Downloads\Bubble_sheet\zipped\1\ID2Q13CH3-20220106T190109Z-001\ID2Q13CH3\6c7802ca-050d-40a6-83e0-27673a80c62d.jpg"]

# gradePapers(paperPaths, modelAnswers)


