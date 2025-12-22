import cv2 as cv
import pandas as pd
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from Module1GradesSheet.src.Extractions.cellsExtraction import detect_table_cells
from Module1GradesSheet.src.Extractions.tableExtraction import extractTable
from Module1GradesSheet.src.Recognition.Names.NameRecognition import name_recognizer
from Module1GradesSheet.src.Recognition.NumberRecognition.NumberRecognition import number_recognizer
from Module1GradesSheet.src.Recognition.Symbols.SymbolRecognition import symbol_recognizer

class FileHandler:
    def __init__(self):
        self.name_engine = name_recognizer
        self.number_engine = number_recognizer
        self.symbol_engine = symbol_recognizer
        
        self.printed_method = os.getenv("PRINTED_NUMBER_RECOGNITION_METHOD", "TRADITIONAL")
        self.written_method = os.getenv("WRITTEN_NUMBER_RECOGNITION_METHOD", "ALREADY_MADE_OCR")
        self.segmentation_method = os.getenv("SEGMENTATION_NUMBER_RECOGNITION_METHOD", "TRADITIONAL")
        self.symbol_method = os.getenv("SYMBOL_RECOGNITION_METHOD", "TRADITIONAL")
        
        print(f"--- FileHandler Ready ---")

    def process_image_to_excel(self, image_path, column_config, output_name="results.xlsx"):
        print(f"--- Extracting Table from {image_path} ---")
        table_img = extractTable(image_path)
        rows = detect_table_cells(table_img)

        data_rows = rows[1:]
        data_results = []

        for i, row in enumerate(data_rows):
            row_data = {}
            print(f"Processing Data Row {i+2}...")

            for j, cell_box in enumerate(row):
                if j >= len(column_config): break
                
                x, y, w, h = cell_box
                cell_roi = table_img[y:y+h, x:x+w]
                
                # Get configuration for this specific column
                config = column_config[j]
                col_name = config['name']
                col_type = config['type']
                # Get expected length if it exists, otherwise None
                expected_len = config.get('len', None)

                # Pass the expected length to the routing function
                result, bg_color = self._route_cell(cell_roi, col_type, expected_len)
                
                row_data[col_name] = result
                row_data[f"{col_name}_bg"] = bg_color
            
            data_results.append(row_data)

        self._save_to_styled_excel(data_results, column_config, output_name)

    def _route_cell(self, cell_roi, col_type, expected_len=None):
        """Routes cell to engine, passing the expected digit length."""
        
        # 1. Printed Student ID
        if col_type == "Number":
            val = self.number_engine.predict_id_from_cell(
                cell_roi, 
                recognition_method=self.printed_method,
                segmentation_method=self.segmentation_method,
                expected_digits=expected_len
            )
            return val, "WHITE"
        
        # 2. Handwritten/Written Number
        elif col_type == "Written Number":
            val = self.number_engine.predict_id_from_cell(
                cell_roi, 
                recognition_method=self.written_method,
                segmentation_method=self.segmentation_method,
                expected_digits=expected_len
            )
            return val, "WHITE"
        
        # 3. Names
        elif col_type == "Arabic Name":
            val = self.name_engine.recognize_student_name(cell_roi, "ar")
            return val, "WHITE"
        
        elif col_type == "English Name":
            val = self.name_engine.recognize_student_name(cell_roi, "en")
            return val, "WHITE"
        
        elif col_type == "Symbol":
            val = self.symbol_engine.predict(cell_roi, override_method=self.symbol_method)
            colour = "WHITE"
            if val == -1:
                colour = "RED"
                val = ""

            return val, colour

        return "", "WHITE"

    def _save_to_styled_excel(self, data, config, output_name):
        wb = Workbook()
        ws = wb.active
        headers = [c['name'] for c in config]
        ws.append(headers)

        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        for row_idx, entry in enumerate(data, start=2):
            for col_idx, col_conf in enumerate(config, start=1):
                name = col_conf['name']
                val = entry.get(name, "")
                bg = entry.get(f"{name}_bg", "WHITE")
                
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if bg == "RED":
                    cell.fill = red_fill

        if not output_name.endswith('.xlsx'):
            output_name = output_name.rsplit('.', 1)[0] + '.xlsx'

        wb.save(output_name)
        print(f"Successfully saved results to {output_name}")

# ---------------------- SINGLETON ----------------------
fileHandler = FileHandler()
